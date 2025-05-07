from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from collections import defaultdict
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta'

DATES = [
    "27/05/2025 (Terça-feira) – 14h às 16h",
    "29/05/2025 (Quinta-feira) – 10h às 12h",
    "03/06/2025 (Terça-feira) – 14h às 16h",
    "05/06/2025 (Quinta-feira) – 10h às 12h",
    "10/06/2025 (Terça-feira) – 14h às 16h",
    "12/06/2025 (Quinta-feira) – 10h às 12h",
    "24/06/2025 (Terça-feira) – 14h às 16h"
]

MAX_PER_DATE = 30
EXCEL_FILE = 'cadastros.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cadastros"
        ws.append(["Nome Completo", "Email", "Setor", "Gestor", "Data do Treinamento"])
        wb.save(EXCEL_FILE)

def add_to_excel(name, email, sector, manager, date):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([name, email, sector, manager, date])
    wb.save(EXCEL_FILE)

def count_registrations():
    # Conta quantos cadastros por data no Excel
    if not os.path.exists(EXCEL_FILE):
        return defaultdict(int)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    counts = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[4]:
            counts[row[4]] += 1
    return counts

@app.route('/', methods=['GET', 'POST'])
def index():
    init_excel()
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        sector = request.form['sector']
        manager = request.form['manager']
        date = request.form['date']

        counts = count_registrations()

        if not all([name, email, sector, manager, date]):
            flash('Por favor, preencha todos os campos.', 'danger')
        elif counts[date] >= MAX_PER_DATE:
            flash('Esta data já está lotada. Por favor, escolha outra.', 'danger')
        else:
            add_to_excel(name, email, sector, manager, date)
            flash('Inscrição realizada com sucesso!', 'success')
            return redirect(url_for('index'))

    counts = count_registrations()
    date_status = []
    for d in DATES:
        if counts[d] >= MAX_PER_DATE:
            date_status.append((d, True))
        else:
            date_status.append((d, False))

    return render_template('index.html', date_status=date_status)

@app.route('/baixar', methods=['POST'])
def baixar():
    senha = request.form.get('senha')
    if senha == '123321':
        init_excel()
        return send_file(EXCEL_FILE, as_attachment=True)
    else:
        return jsonify({'error': 'Senha incorreta!'}), 401

if __name__ == '__main__':
    app.run(debug=True)
